import polars as pl
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

def export_report_to_csv(data, filename='report.csv', include_summary=False, summary_data=None):
    """
    Export a list of dicts or a polars DataFrame to CSV with formatting.
    
    Args:
        data: list[dict] | pl.DataFrame
        filename: output filename
        include_summary: whether to include summary rows
        summary_data: dict with summary information (totals, counts, etc.)
    """
    if not isinstance(data, pl.DataFrame):
        df = pl.DataFrame(data)
    else:
        df = data
    
    # Format numeric columns
    for col in df.columns:
        if df[col].dtype in [pl.Float32, pl.Float64]:
            df = df.with_columns(pl.col(col).cast(pl.Utf8).map_elements(lambda x: f"{float(x) if x else 0:,.2f}" if isinstance(x, (int, float)) else x, return_dtype=pl.Utf8))
    
    csv_bytes = df.write_csv()
    
    # Append summary if provided
    if include_summary and summary_data:
        summary_csv = "\n\nSummary\n"
        for key, value in summary_data.items():
            if isinstance(value, float):
                summary_csv += f"{key},{value:,.2f}\n"
            else:
                summary_csv += f"{key},{value}\n"
        csv_bytes = csv_bytes + summary_csv.encode('utf-8')
    
    response = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_report_to_xlsx(data, filename='report.xlsx', title=None, company=None, include_summary=False, summary_data=None):
    """
    Export a list of dicts or polars DataFrame to Excel with professional formatting.
    
    Args:
        data: list[dict] | pl.DataFrame
        filename: output filename
        title: optional document title
        company: optional dict with company details
        include_summary: whether to include summary row
        summary_data: dict with summary information
    
    Returns:
        HttpResponse with Excel file
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not available, falling back to CSV export")
        return export_report_to_csv(data, filename.replace('.xlsx', '.csv'))
    
    # Normalize to polars DataFrame
    if not isinstance(data, pl.DataFrame):
        try:
            df = pl.DataFrame(data or [])
        except Exception:
            df = pl.DataFrame([])
    else:
        df = data
    
    if df.is_empty():
        df = pl.DataFrame([])
    
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Report"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    company_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Header section
    row = 1
    
    # Company details
    if company:
        company_name = company.get('name', 'BengoERP')
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = company_name
        cell.font = title_font
        row += 1
        
        company_details = []
        for key in ['address', 'email', 'phone']:
            val = company.get(key)
            if val:
                company_details.append(str(val))
        
        if company_details:
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = ' | '.join(company_details)
            cell.font = company_font
            row += 1
    
    # Title and date
    if title:
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = title_font
        row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    cell.font = company_font
    row += 2  # Add spacing
    
    # Column headers
    if not df.is_empty():
        columns = df.columns
        header_row = row
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Data rows
        row += 1
        for data_row in df.iter_rows(named=True):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_idx)
                value = data_row.get(col_name)
                cell.value = _format_excel_cell(value)
                cell.border = border
                
                # Right-align numbers
                if isinstance(value, (int, float)):
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            row += 1
        
        # Summary row if provided
        if include_summary and summary_data:
            row += 1
            summary_col = 1
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = "TOTALS"
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            
            for col_idx, col_name in enumerate(columns, 1):
                if col_name in summary_data:
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = summary_data[col_name]
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                    cell.number_format = '#,##0.00' if isinstance(summary_data[col_name], float) else '#,##0'
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for data_row in df.iter_rows(named=True):
                cell_len = len(str(data_row.get(col_name) or ''))
                max_length = max(max_length, cell_len)
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Generate file
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_report_to_pdf(data, filename='report.pdf', title=None, company=None, footer_text=None):
    """
    Export a list of dicts or polars DataFrame to a tabular PDF with
    a consistent header and footer using reportlab.

    Args:
        data: list[dict] | pl.DataFrame
        filename: output filename
        title: optional document title (string)
        company: optional dict with keys {name, address, email, phone, registration}
        footer_text: optional footer override
    """
    # Normalize to polars DataFrame
    if not isinstance(data, pl.DataFrame):
        try:
            df = pl.DataFrame(data or [])
        except Exception:
            df = pl.DataFrame([])
    else:
        df = data

    # Derive columns
    if isinstance(data, list) and data and isinstance(data[0], dict):
        first_keys = list(data[0].keys())
        extra_keys = []
        for row in data:
            for k in row.keys():
                if k not in first_keys and k not in extra_keys:
                    extra_keys.append(k)
        columns = first_keys + extra_keys
        for col in columns:
            if col not in df.columns:
                df = df.with_columns(pl.lit(None).alias(col))
        df = df.select(columns)
    else:
        columns = df.columns if not df.is_empty() else []

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.6*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Header block with company details
    header_title = title or 'Report'
    company_name = (company or {}).get('name') if company else None
    company_lines = []
    if company:
        for key in ['address', 'email', 'phone', 'registration']:
            val = company.get(key)
            if val:
                company_lines.append(str(val))

    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1a1a'), spaceAfter=6)
    subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Normal'], fontSize=9, textColor=colors.grey)

    if company_name:
        elements.append(Paragraph(company_name, title_style))
        if company_lines:
            elements.append(Paragraph(' | '.join(company_lines), subtitle_style))
        elements.append(Spacer(1, 0.15*inch))

    elements.append(Paragraph(header_title, ParagraphStyle('DocTitle', parent=styles['Heading2'], spaceAfter=8)))
    elements.append(Paragraph(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    # Table data
    if columns:
        table_data = [columns]
        for row in df.iter_rows(named=True):
            table_data.append([_format_cell(row.get(col)) for col in columns])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

    # Footer note
    footer = footer_text or 'Generated by BengoERP | Confidential'
    elements.append(Spacer(1, 0.15*inch))
    elements.append(Paragraph(footer, ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _format_cell(value):
    """Format a cell value for display."""
    if value is None:
        return ''
    if isinstance(value, float):
        try:
            return f"{value:,.2f}"
        except Exception:
            return str(value)
    return str(value)


def _format_excel_cell(value):
    """Format a cell value for Excel."""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return value
    return str(value)


def get_company_details_from_request(request):
    """
    Extract company details from request session or context.
    Returns dict with: name, address, email, phone, registration, logo_url
    """
    try:
        from business.models import Bussiness, Branch
        
        # Get from context if available
        if hasattr(request, 'business_id'):
            business_id = request.business_id
        elif hasattr(request, 'branch_id'):
            from business.models import Branch
            branch = Branch.objects.get(id=request.branch_id)
            business_id = branch.business_id
        else:
            return None
        
        business = Bussiness.objects.get(id=business_id)
        main_branch = Branch.objects.filter(business=business).first()
        
        if main_branch:
            return {
                'name': business.name,
                'address': main_branch.location.city if main_branch.location else 'Kenya',
                'email': main_branch.email or business.email,
                'phone': main_branch.contact_number or business.contact_number,
                'registration': business.kra_number or 'Not configured',
                'logo_url': business.logo.url if business.logo else None
            }
    except Exception as e:
        logger.warning(f"Could not fetch company details: {str(e)}")
    
    return None


# Availability check
def has_excel_support():
    """Check if Excel export is available."""
    return OPENPYXL_AVAILABLE 